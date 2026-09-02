"""Generation et lecture des devis Excel, FIDELE au template (logo, styles, pied de page).

IMPORTANT : on n'utilise PAS openpyxl pour ECRIRE le fichier final, car openpyxl supprime
le logo (image ancree) et le tableau stylise. On fait de la "chirurgie XML" sur une copie
du template : on clone la ligne du tableau des produits, on decale les lignes suivantes, et
on laisse tout le reste (logo, styles, formules de total, pied de page) intact.

openpyxl sert uniquement a la LECTURE (scan d'un .xlsx cree/modifie a la main).
"""
import math
import os
import re
import shutil
import zipfile
from datetime import date, datetime
from pathlib import Path
from xml.sax.saxutils import escape

import openpyxl

from .config import settings
from .schemas import DevisData, LigneDevis

# --- Constantes liees a la structure du template ---
HEADER_ROW = 17
BODY_ROW = 18          # 1ere (et unique) ligne de corps du tableau dans le template
ROW_FORMULA = "Tableau1[[#This Row],[Prix Unit HT]]*Tableau1[[#This Row],[Quantité]]"
# Estimation de hauteur de ligne pour le texte multiligne (colonne Produit large).
_CHARS_PER_LINE = 55
_LINE_PT = 16
_PAD_PT = 8


# ============================ ECRITURE ============================

def _serial(d: date) -> int:
    """Numero de serie Excel d'une date."""
    return (d - date(1899, 12, 30)).days


def _est_height(text: str) -> float:
    lines = 0
    for seg in (text or "").split("\n"):
        lines += max(1, math.ceil(len(seg) / _CHARS_PER_LINE))
    return max(16.0, lines * _LINE_PT + _PAD_PT)


def _num(v) -> str:
    """Formate un nombre pour le XML (entier sans .0)."""
    if v is None:
        return "0"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _body_row_xml(r: int, ligne: LigneDevis) -> str:
    h = _est_height(ligne.produit)
    produit = escape(ligne.produit or "")
    tva_frac = (ligne.tva_pct or 0) / 100.0
    return (
        f'<row r="{r}" spans="1:5" ht="{h:.0f}" customHeight="1" x14ac:dyDescent="0.2">'
        f'<c r="A{r}" s="13" t="inlineStr"><is><t xml:space="preserve">{produit}</t></is></c>'
        f'<c r="B{r}"><v>{_num(ligne.quantite)}</v></c>'
        f'<c r="C{r}" s="3"><v>{_num(ligne.prix_unit_ht)}</v></c>'
        f'<c r="D{r}" s="11"><v>{tva_frac}</v></c>'
        f'<c r="E{r}" s="3"><f>{ROW_FORMULA}</f></c>'
        f'</row>'
    )


def _shift_row(row_xml: str, delta: int) -> str:
    row_xml = re.sub(r'(<row [^>]*\br=")(\d+)(")',
                     lambda m: m.group(1) + str(int(m.group(2)) + delta) + m.group(3), row_xml)
    row_xml = re.sub(r'(r="[A-E])(\d+)(")',
                     lambda m: m.group(1) + str(int(m.group(2)) + delta) + m.group(3), row_xml)
    return row_xml


def _build_sheet(xml: str, lignes: list[LigneDevis]) -> str:
    n = max(len(lignes), 1)
    delta = n - 1
    sd = re.search(r'<sheetData>(.*)</sheetData>', xml, re.S)
    rows = re.findall(r'<row\b[^>]*?(?:/>|>.*?</row>)', sd.group(1), re.S)

    out = []
    for rx in rows:
        rn = int(re.search(r'<row [^>]*\br="(\d+)"', rx).group(1))
        if rn < BODY_ROW:
            out.append(rx)
        elif rn == BODY_ROW:
            for i in range(n):
                ligne = lignes[i] if i < len(lignes) else LigneDevis()
                out.append(_body_row_xml(BODY_ROW + i, ligne))
        else:
            out.append(_shift_row(rx, delta) if delta else rx)

    xml = xml[:sd.start()] + "<sheetData>" + "".join(out) + "</sheetData>" + xml[sd.end():]
    xml = re.sub(r'<dimension ref="[^"]*"/>', f'<dimension ref="A2:E{26 + delta}"/>', xml)
    # Force le recalcul en retirant les valeurs en cache des cellules a formule.
    xml = re.sub(r'(<c\b[^>]*?)\s+t="(?:e|str)"([^>]*>\s*<f)', r'\1\2', xml)
    xml = re.sub(r'(</f>)<v>.*?</v>', r'\1', xml)
    return xml


def _set_cell_text(xml: str, coord: str, text: str, style: str) -> str:
    """Ecrit une chaine dans une cellule de la ligne d'entete (au-dessus du tableau)."""
    cell = (f'<c r="{coord}" s="{style}" t="inlineStr"><is>'
            f'<t xml:space="preserve">{escape(text)}</t></is></c>' if text
            else f'<c r="{coord}" s="{style}"/>')
    if re.search(rf'<c r="{coord}"[^>]*/>', xml):
        return re.sub(rf'<c r="{coord}"[^>]*/>', cell, xml, count=1)
    if re.search(rf'<c r="{coord}"', xml):
        return re.sub(rf'<c r="{coord}".*?</c>', cell, xml, count=1, flags=re.S)
    row = int(coord[1:])
    if re.search(rf'<row r="{row}"[^>]*>', xml):
        return re.sub(rf'(<row r="{row}"[^>]*>)', rf'\1{cell}', xml, count=1)
    # Ligne absente : on l'insere a sa position (avant la 1ere ligne de numero superieur),
    # pour garder l'ordre croissant exige par le format xlsx.
    new_row = f'<row r="{row}" spans="1:5" x14ac:dyDescent="0.2">{cell}</row>'
    for existing in sorted(set(int(m) for m in re.findall(r'<row r="(\d+)"', xml))):
        if existing > row:
            return re.sub(rf'(<row r="{existing}")', rf'{new_row}\1', xml, count=1)
    return xml


def _set_client(xml: str, data: DevisData) -> str:
    # Bloc client aligne a droite (style s=1), cellules E10 / E11 / E12.
    xml = _set_cell_text(xml, "E10", data.client_nom, "1")
    xml = _set_cell_text(xml, "E11", data.client_tel, "1")
    xml = _set_cell_text(xml, "E12", data.client_email, "1")
    return xml


def _set_dates(xml: str, date_devis: date, date_prestation) -> str:
    # E2 = date de devis : remplace =TODAY() par une date figee.
    xml = re.sub(r'<c r="E2".*?</c>',
                 f'<c r="E2" s="12"><v>{_serial(date_devis)}</v></c>', xml, count=1, flags=re.S)
    # B16 = date de prestation en TEXTE aligne a gauche (colonne trop etroite pour une vraie date).
    if date_prestation:
        txt = "   " + date_prestation.strftime("%d/%m/%Y")  # petit decalage visuel apres le libelle
        cell = f'<c r="B16" s="1" t="inlineStr"><is><t xml:space="preserve">{txt}</t></is></c>'
        if re.search(r'<c r="B16"[^>]*/>', xml):
            xml = re.sub(r'<c r="B16"[^>]*/>', cell, xml, count=1)
        elif re.search(r'<c r="B16"', xml):
            xml = re.sub(r'<c r="B16".*?</c>', cell, xml, count=1, flags=re.S)
        else:
            xml = re.sub(r'(<row r="16"[^>]*>)', rf'\1{cell}', xml, count=1)
    return xml


def write_devis(data: DevisData, dest_path: Path) -> Path:
    """Genere un .xlsx fidele au template a partir des donnees structurees."""
    dest_path = Path(dest_path)
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    work = dest_path.with_suffix(".work")
    if work.exists():
        shutil.rmtree(work)
    work.mkdir(parents=True)

    with zipfile.ZipFile(settings.template_file) as z:
        z.extractall(work)

    sheet_p = work / "xl/worksheets/sheet1.xml"
    xml = sheet_p.read_text(encoding="utf-8")
    xml = _build_sheet(xml, data.lignes)
    xml = _set_dates(xml, data.date_devis, data.date_prestation)
    xml = _set_client(xml, data)
    sheet_p.write_text(xml, encoding="utf-8")

    # Force le recalcul complet a l'ouverture.
    wb_p = work / "xl/workbook.xml"
    wbx = wb_p.read_text(encoding="utf-8")
    if "fullCalcOnLoad" not in wbx:
        wbx = re.sub(r'<calcPr ', '<calcPr fullCalcOnLoad="1" ', wbx, count=1)
    wb_p.write_text(wbx, encoding="utf-8")

    # Etend la reference du tableau au nombre de lignes.
    tbl_p = work / "xl/tables/table1.xml"
    n = max(len(data.lignes), 1)
    txml = tbl_p.read_text(encoding="utf-8")
    txml = re.sub(r'ref="A17:E\d+"', f'ref="A17:E{HEADER_ROW + n}"', txml)
    tbl_p.write_text(txml, encoding="utf-8")

    # Supprime calcChain (recalcule automatiquement, evite les avertissements Excel).
    cc = work / "xl/calcChain.xml"
    if cc.exists():
        cc.unlink()
        ct_p = work / "[Content_Types].xml"
        ct_p.write_text(re.sub(r'<Override PartName="/xl/calcChain.xml"[^>]*/>', '',
                               ct_p.read_text(encoding="utf-8")), encoding="utf-8")
        wr_p = work / "xl/_rels/workbook.xml.rels"
        wr_p.write_text(re.sub(r'<Relationship [^>]*calcChain[^>]*/>', '',
                               wr_p.read_text(encoding="utf-8")), encoding="utf-8")

    if dest_path.exists():
        dest_path.unlink()
    with zipfile.ZipFile(dest_path, "w", zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(work):
            for f in files:
                full = Path(root) / f
                z.write(full, full.relative_to(work).as_posix())
    shutil.rmtree(work)
    return dest_path


# ============================ LECTURE (scan) ============================

def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(" ", "").replace(" ", "").replace(",", "."))
    except ValueError:
        return 0.0


def _parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str) and v.strip():
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                continue
    return None


def read_devis(src_path: Path) -> DevisData:
    """Relit un .xlsx (potentiellement modifie a la main) vers DevisData."""
    wb = openpyxl.load_workbook(src_path, data_only=True)
    ws = wb.active

    date_devis = _parse_date(ws["E2"].value) or date.today()
    date_prestation = _parse_date(ws["B16"].value)
    client_nom = str(ws["E10"].value or "").strip()
    client_tel = str(ws["E11"].value or "").strip()
    client_email = str(ws["E12"].value or "").strip()

    # Determine la derniere ligne du tableau via sa reference (sinon on s'arrete au 1er vide).
    last_row = BODY_ROW
    if "Tableau1" in ws.tables:
        m = re.search(r':E(\d+)', ws.tables["Tableau1"].ref)
        if m:
            last_row = int(m.group(1))

    lignes: list[LigneDevis] = []
    for r in range(BODY_ROW, last_row + 1):
        produit = ws.cell(row=r, column=1).value
        prix = _to_float(ws.cell(row=r, column=3).value)
        qte = _to_float(ws.cell(row=r, column=2).value)
        if (produit is None or str(produit).strip() == "") and prix == 0.0:
            continue
        lignes.append(LigneDevis(
            produit=str(produit or ""),
            quantite=qte,
            prix_unit_ht=prix,
            tva_pct=round(_to_float(ws.cell(row=r, column=4).value) * 100, 2),
        ))

    return DevisData(
        date_devis=date_devis, date_prestation=date_prestation,
        client_nom=client_nom, client_tel=client_tel, client_email=client_email,
        lignes=lignes,
    )
